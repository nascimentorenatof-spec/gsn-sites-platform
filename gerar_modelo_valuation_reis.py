from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


OUTPUT = "Modelo_Valuation_Fabrica_Reis_Lelooks_Piratas.xlsx"


def money_fmt():
    return 'R$ #,##0.0;[Red](R$ #,##0.0);-'


def pct_fmt():
    return '0.0%'


def multiple_fmt():
    return '0.0x'


def style_sheet(ws):
    thin = Side(style="thin", color="D9DEE7")
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center")
            cell.border = Border(bottom=thin)
            if cell.row == 1:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1F4E79")
    ws.sheet_view.showGridLines = False


def set_widths(ws, widths):
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def section(ws, row, title):
    ws.cell(row, 1, title)
    ws.cell(row, 1).font = Font(bold=True, color="FFFFFF")
    ws.cell(row, 1).fill = PatternFill("solid", fgColor="5B9BD5")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)


def main():
    wb = Workbook()
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True

    ws_inputs = wb.active
    ws_inputs.title = "Inputs"
    ws_fcd = wb.create_sheet("FCD")
    ws_mult = wb.create_sheet("Multiplos")
    ws_pat = wb.create_sheet("Patrimonial")
    ws_sens = wb.create_sheet("Sensibilidade")
    ws_resumo = wb.create_sheet("Resumo")
    ws_fontes = wb.create_sheet("Fontes")

    # Inputs
    ws_inputs.append(["Modelo de Valuation - Fabrica Reis / marcas Lelooks e Piratas for men", "", "", "", "", "", "", ""])
    ws_inputs.merge_cells("A1:H1")
    ws_inputs["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws_inputs["A1"].fill = PatternFill("solid", fgColor="1F4E79")

    section(ws_inputs, 3, "Premissas operacionais editaveis")
    rows = [
        ("Receita base 2025", 18_000_000, "Receita liquida anual normalizada da fabrica", "R$"),
        ("Participacao Lelooks", 0.58, "Mix de receita feminina", "%"),
        ("Participacao Piratas for men", "=1-B5", "Mix de receita masculina", "%"),
        ("Crescimento receita ano 1", 0.12, "Expansao de vendas", "%"),
        ("Crescimento receita ano 2", 0.10, "Expansao de vendas", "%"),
        ("Crescimento receita ano 3", 0.08, "Expansao de vendas", "%"),
        ("Crescimento receita ano 4", 0.07, "Expansao de vendas", "%"),
        ("Crescimento receita ano 5", 0.06, "Expansao de vendas", "%"),
        ("Margem bruta", 0.48, "Receita menos CPV", "%"),
        ("Despesas comerciais/administrativas", 0.28, "SG&A sobre receita", "%"),
        ("Ajustes de EBITDA", 0, "Ajustes recorrentes: pro-labore, alugueis, one-offs", "R$"),
        ("Depreciacao / receita", 0.03, "Depreciacao e amortizacao", "%"),
        ("Capex / receita", 0.04, "Reposicao e expansao de maquinas", "%"),
        ("Capital de giro / incremento receita", 0.12, "Necessidade incremental", "%"),
        ("Aliquota IR/CSLL efetiva", 0.24, "Taxa efetiva estimada", "%"),
    ]
    start = 4
    for idx, (label, value, note, unit) in enumerate(rows, start):
        ws_inputs.cell(idx, 1, label)
        ws_inputs.cell(idx, 2, value)
        ws_inputs.cell(idx, 3, note)
        ws_inputs.cell(idx, 4, unit)

    section(ws_inputs, 21, "Taxa de desconto: build-up de risco setorial")
    risk_rows = [
        ("Taxa livre de risco / CDI longo prazo", 0.105, "Base Brasil nominal", "%"),
        ("Premio de risco de mercado", 0.065, "Risco de equity", "%"),
        ("Beta desalavancado moda/varejo", 0.90, "Sensibilidade setorial", "x"),
        ("Premio tamanho/empresa fechada", 0.035, "Pequena empresa e liquidez", "%"),
        ("Premio concentracao de clientes/canais", 0.020, "Dependencia comercial", "%"),
        ("Premio moda/estoque/sazonalidade", 0.020, "Colecoes, ruptura e markdowns", "%"),
        ("Custo de capital proprio", "=B22+B23*B24+B25+B26+B27", "CAPM ajustado", "%"),
        ("Custo da divida antes IR", 0.145, "Custo bancario estimado", "%"),
        ("Peso equity", 0.75, "Estrutura alvo", "%"),
        ("Peso divida", "=1-B30", "Estrutura alvo", "%"),
        ("WACC nominal", "=B30*B28+B31*B29*(1-B18)", "Taxa de desconto FCD", "%"),
        ("Crescimento perpetuo", 0.045, "Terminal nominal conservador", "%"),
    ]
    for idx, (label, value, note, unit) in enumerate(risk_rows, 22):
        ws_inputs.cell(idx, 1, label)
        ws_inputs.cell(idx, 2, value)
        ws_inputs.cell(idx, 3, note)
        ws_inputs.cell(idx, 4, unit)

    section(ws_inputs, 36, "Pesos do valor final")
    weights = [
        ("Peso FCD", 0.55),
        ("Peso multiplos", 0.25),
        ("Peso patrimonial", 0.20),
        ("Total", "=SUM(B37:B39)"),
    ]
    for idx, (label, value) in enumerate(weights, 37):
        ws_inputs.cell(idx, 1, label)
        ws_inputs.cell(idx, 2, value)

    for row in range(4, 41):
        if ws_inputs.cell(row, 4).value == "%":
            ws_inputs.cell(row, 2).number_format = pct_fmt()
        elif ws_inputs.cell(row, 4).value == "R$":
            ws_inputs.cell(row, 2).number_format = money_fmt()
        elif ws_inputs.cell(row, 4).value == "x":
            ws_inputs.cell(row, 2).number_format = multiple_fmt()
    for row in range(37, 41):
        ws_inputs.cell(row, 2).number_format = pct_fmt()
    set_widths(ws_inputs, {"A": 34, "B": 18, "C": 52, "D": 10})
    style_sheet(ws_inputs)

    # FCD
    ws_fcd.append(["Fluxo de Caixa Descontado", "2025 base", "Ano 1", "Ano 2", "Ano 3", "Ano 4", "Ano 5"])
    fcd_rows = [
        ("Receita liquida", "=Inputs!B4", "=B2*(1+Inputs!B7)", "=C2*(1+Inputs!B8)", "=D2*(1+Inputs!B9)", "=E2*(1+Inputs!B10)", "=F2*(1+Inputs!B11)"),
        ("Lelooks", "=B2*Inputs!B5", "=C2*Inputs!B5", "=D2*Inputs!B5", "=E2*Inputs!B5", "=F2*Inputs!B5", "=G2*Inputs!B5"),
        ("Piratas for men", "=B2*Inputs!B6", "=C2*Inputs!B6", "=D2*Inputs!B6", "=E2*Inputs!B6", "=F2*Inputs!B6", "=G2*Inputs!B6"),
        ("CPV", "=-B2*(1-Inputs!B12)", "=-C2*(1-Inputs!B12)", "=-D2*(1-Inputs!B12)", "=-E2*(1-Inputs!B12)", "=-F2*(1-Inputs!B12)", "=-G2*(1-Inputs!B12)"),
        ("Lucro bruto", "=B2+B5", "=C2+C5", "=D2+D5", "=E2+E5", "=F2+F5", "=G2+G5"),
        ("Despesas comerciais/adm.", "=-B2*Inputs!B13", "=-C2*Inputs!B13", "=-D2*Inputs!B13", "=-E2*Inputs!B13", "=-F2*Inputs!B13", "=-G2*Inputs!B13"),
        ("EBITDA ajustado", "=B6+B7+Inputs!B14", "=C6+C7+Inputs!B14", "=D6+D7+Inputs!B14", "=E6+E7+Inputs!B14", "=F6+F7+Inputs!B14", "=G6+G7+Inputs!B14"),
        ("Margem EBITDA", "=B8/B2", "=C8/C2", "=D8/D2", "=E8/E2", "=F8/F2", "=G8/G2"),
        ("Depreciacao", "=-B2*Inputs!B15", "=-C2*Inputs!B15", "=-D2*Inputs!B15", "=-E2*Inputs!B15", "=-F2*Inputs!B15", "=-G2*Inputs!B15"),
        ("EBIT", "=B8+B10", "=C8+C10", "=D8+D10", "=E8+E10", "=F8+F10", "=G8+G10"),
        ("Impostos", "=-MAX(0,B11)*Inputs!B18", "=-MAX(0,C11)*Inputs!B18", "=-MAX(0,D11)*Inputs!B18", "=-MAX(0,E11)*Inputs!B18", "=-MAX(0,F11)*Inputs!B18", "=-MAX(0,G11)*Inputs!B18"),
        ("NOPAT", "=B11+B12", "=C11+C12", "=D11+D12", "=E11+E12", "=F11+F12", "=G11+G12"),
        ("+ Depreciacao", "=-B10", "=-C10", "=-D10", "=-E10", "=-F10", "=-G10"),
        ("- Capex", "=-B2*Inputs!B16", "=-C2*Inputs!B16", "=-D2*Inputs!B16", "=-E2*Inputs!B16", "=-F2*Inputs!B16", "=-G2*Inputs!B16"),
        ("- Capital de giro incremental", 0, "=-(C2-B2)*Inputs!B17", "=-(D2-C2)*Inputs!B17", "=-(E2-D2)*Inputs!B17", "=-(F2-E2)*Inputs!B17", "=-(G2-F2)*Inputs!B17"),
        ("Fluxo de caixa livre", "=SUM(B13:B16)", "=SUM(C13:C16)", "=SUM(D13:D16)", "=SUM(E13:E16)", "=SUM(F13:F16)", "=SUM(G13:G16)"),
        ("Fator desconto", "", "=1/(1+Inputs!B32)^1", "=1/(1+Inputs!B32)^2", "=1/(1+Inputs!B32)^3", "=1/(1+Inputs!B32)^4", "=1/(1+Inputs!B32)^5"),
        ("VP FCL", "", "=C17*C18", "=D17*D18", "=E17*E18", "=F17*F18", "=G17*G18"),
        ("Valor terminal", "", "", "", "", "", "=G17*(1+Inputs!B33)/(Inputs!B32-Inputs!B33)"),
        ("VP valor terminal", "", "", "", "", "", "=G20*G18"),
        ("Enterprise value FCD", "", "", "", "", "", "=SUM(C19:G19)+G21"),
    ]
    for r in fcd_rows:
        ws_fcd.append(list(r))
    for row in range(2, 23):
        for col in range(2, 8):
            ws_fcd.cell(row, col).number_format = pct_fmt() if row in [9, 18] else money_fmt()
    set_widths(ws_fcd, {"A": 32, "B": 16, "C": 16, "D": 16, "E": 16, "F": 16, "G": 16})
    style_sheet(ws_fcd)

    chart = LineChart()
    chart.title = "Receita e EBITDA projetados"
    chart.y_axis.title = "R$"
    chart.x_axis.title = "Ano"
    chart.add_data(Reference(ws_fcd, min_col=3, max_col=7, min_row=2, max_row=2), from_rows=True, titles_from_data=False)
    chart.add_data(Reference(ws_fcd, min_col=3, max_col=7, min_row=8, max_row=8), from_rows=True, titles_from_data=False)
    chart.set_categories(Reference(ws_fcd, min_col=3, max_col=7, min_row=1, max_row=1))
    chart.height = 7
    chart.width = 16
    ws_fcd.add_chart(chart, "I2")

    # Multiplos
    ws_mult.append(["Pares brasileiros de moda/textil", "Ticker", "Receita LTM / anual", "EBITDA LTM / anual", "EV/Receita", "EV/EBITDA", "P/L", "Observacao"])
    peers = [
        ("Lojas Renner", "LREN3", 13_838_200_000, 3_200_000_000, 1.1, 5.3, 13.0, "Varejo moda omnichannel"),
        ("Riachuelo / Guararapes", "RIAA3", 10_500_000_000, 1_800_000_000, 0.7, 4.2, 9.5, "Varejo moda + financeiro"),
        ("C&A Brasil", "CEAB3", 7_980_000_000, 1_240_000_000, 0.6, 4.0, 8.5, "Varejo moda"),
        ("Azzas 2154", "AZZA3", 11_800_000_000, 1_940_000_000, 1.0, 7.0, 14.0, "Grupo de marcas"),
        ("Vulcabras", "VULC3", 4_200_000_000, 763_100_000, 1.3, 6.2, 10.5, "Calcados e vestuario esportivo"),
    ]
    for peer in peers:
        ws_mult.append(list(peer))
    ws_mult.append(["Mediana", "", "=MEDIAN(C2:C6)", "=MEDIAN(D2:D6)", "=MEDIAN(E2:E6)", "=MEDIAN(F2:F6)", "=MEDIAN(G2:G6)", ""])
    ws_mult.append(["Desconto empresa fechada/tamanho", "", "", "", 0.30, 0.30, 0.30, "Ajustavel"])
    ws_mult.append(["Multiplo aplicado", "", "", "", "=E7*(1-E8)", "=F7*(1-F8)", "=G7*(1-G8)", ""])
    ws_mult.append(["Receita Reis Ano 1", "", "=FCD!C2", "", "", "", "", ""])
    ws_mult.append(["EBITDA Reis Ano 1", "", "", "=FCD!C8", "", "", "", ""])
    ws_mult.append(["Lucro liquido proxy", "", "", "=FCD!C13", "", "", "", "NOPAT como proxy"])
    ws_mult.append(["Valor por EV/Receita", "", "", "", "=C10*E9", "", "", ""])
    ws_mult.append(["Valor por EV/EBITDA", "", "", "", "", "=D11*F9", "", ""])
    ws_mult.append(["Valor por P/L", "", "", "", "", "", "=D12*G9", ""])
    ws_mult.append(["Enterprise value por multiplos", "", "", "", "", "", "=AVERAGE(E13,F14,G15)", ""])
    for row in range(2, 17):
        for col in [3, 4, 5, 6, 7]:
            if col in [5, 6, 7] and row <= 9:
                ws_mult.cell(row, col).number_format = pct_fmt() if row == 8 else multiple_fmt()
            else:
                ws_mult.cell(row, col).number_format = money_fmt()
    set_widths(ws_mult, {"A": 34, "B": 12, "C": 18, "D": 18, "E": 14, "F": 14, "G": 14, "H": 34})
    style_sheet(ws_mult)

    # Patrimonial
    ws_pat.append(["Valor Patrimonial Ajustado", "Valor contabil", "Ajuste a mercado", "Valor ajustado", "Observacao"])
    patr_rows = [
        ("Maquinas de corte/costura/acabamento", 2_200_000, 0.10, "=B2*(1+C2)", "Atualizar por laudo/valor de reposicao"),
        ("Instalacoes e benfeitorias", 900_000, 0.00, "=B3*(1+C3)", "Galpao, layout fabril, eletrica"),
        ("Estoque materia-prima e acabado", 3_400_000, -0.08, "=B4*(1+C4)", "Aplicar markdown/obsolescencia"),
        ("Contas a receber", 1_900_000, -0.03, "=B5*(1+C5)", "Risco de inadimplencia"),
        ("Caixa e equivalentes", 450_000, 0.00, "=B6*(1+C6)", ""),
        ("Outros ativos", 250_000, 0.00, "=B7*(1+C7)", ""),
        ("Dividas bancarias", -2_800_000, 0.00, "=B8*(1+C8)", "Valor liquido negativo"),
        ("Fornecedores/impostos/contingencias", -1_200_000, 0.00, "=B9*(1+C9)", "Valor liquido negativo"),
        ("Patrimonio liquido ajustado", "", "", "=SUM(D2:D9)", ""),
    ]
    for r in patr_rows:
        ws_pat.append(list(r))
    ws_pat.append(["", "", "", "", ""])
    ws_pat.append(["Intangiveis de marca", "Receita atribuida", "Royalty rate", "Valor da marca", "Observacao"])
    brand_start = 14
    brand_rows = [
        ("Marca Lelooks", "=FCD!C2*Inputs!B5", 0.025, "=B13*C13*(1-Inputs!B18)/(Inputs!B32-Inputs!B33)", "Metodo relief-from-royalty"),
        ("Marca Piratas for men", "=FCD!C2*Inputs!B6", 0.020, "=B14*C14*(1-Inputs!B18)/(Inputs!B32-Inputs!B33)", "Metodo relief-from-royalty"),
        ("Valor patrimonial + intangiveis", "", "", "=D10+SUM(D13:D14)", ""),
    ]
    for r in brand_rows:
        ws_pat.append(list(r))
    for row in range(2, 17):
        for col in range(2, 5):
            ws_pat.cell(row, col).number_format = pct_fmt() if col == 3 else money_fmt()
    set_widths(ws_pat, {"A": 36, "B": 18, "C": 16, "D": 18, "E": 40})
    style_sheet(ws_pat)

    # Sensibilidade
    ws_sens.append(["Sensibilidade FCD - EV por WACC e crescimento perpetuo"])
    ws_sens.append(["WACC \\ g", 0.025, 0.035, 0.045, 0.055, 0.065])
    waccs = [0.145, 0.160, 0.175, 0.190, 0.205]
    for idx, wacc in enumerate(waccs, 3):
        ws_sens.cell(idx, 1, wacc)
        for col in range(2, 7):
            g_cell = ws_sens.cell(2, col).coordinate
            w_cell = ws_sens.cell(idx, 1).coordinate
            ws_sens.cell(idx, col, f"=SUM(FCD!C19:G19)+FCD!G17*(1+{g_cell})/({w_cell}-{g_cell})*FCD!G18")
            ws_sens.cell(idx, col).number_format = money_fmt()
    ws_sens.append([""])
    ws_sens.append(["Sensibilidade multiplos - EV por margem EBITDA e EV/EBITDA"])
    ws_sens.append(["Margem EBITDA \\ Multiplo", 3.0, 4.0, 5.0, 6.0, 7.0])
    margins = [0.12, 0.16, 0.20, 0.24, 0.28]
    for idx, margin in enumerate(margins, 10):
        ws_sens.cell(idx, 1, margin)
        for col in range(2, 7):
            ws_sens.cell(idx, col, f"=FCD!C2*{ws_sens.cell(idx, 1).coordinate}*{ws_sens.cell(9, col).coordinate}")
            ws_sens.cell(idx, col).number_format = money_fmt()
    for cell in ws_sens["2:2"]:
        cell.number_format = pct_fmt()
    for row in range(3, 8):
        ws_sens.cell(row, 1).number_format = pct_fmt()
    for cell in ws_sens["9:9"]:
        cell.number_format = multiple_fmt()
    for row in range(10, 15):
        ws_sens.cell(row, 1).number_format = pct_fmt()
    for rng in ["B3:F7", "B10:F14"]:
        ws_sens.conditional_formatting.add(
            rng,
            ColorScaleRule(start_type="min", start_color="F8696B", mid_type="percentile", mid_value=50, mid_color="FFEB84", end_type="max", end_color="63BE7B"),
        )
    set_widths(ws_sens, {"A": 28, "B": 16, "C": 16, "D": 16, "E": 16, "F": 16})
    style_sheet(ws_sens)

    sens_chart = LineChart()
    sens_chart.title = "FCD por WACC"
    sens_chart.y_axis.title = "Enterprise value"
    sens_chart.x_axis.title = "WACC"
    sens_chart.add_data(Reference(ws_sens, min_col=4, max_col=4, min_row=2, max_row=7), titles_from_data=True)
    sens_chart.set_categories(Reference(ws_sens, min_col=1, min_row=3, max_row=7))
    sens_chart.height = 7
    sens_chart.width = 16
    ws_sens.add_chart(sens_chart, "H2")

    # Resumo
    ws_resumo.append(["Resumo do Valuation - Fabrica Reis", "", "", ""])
    ws_resumo.merge_cells("A1:D1")
    ws_resumo["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws_resumo["A1"].fill = PatternFill("solid", fgColor="1F4E79")
    summary_rows = [
        ("Enterprise value FCD", "=FCD!G22", "Peso", "=Inputs!B37"),
        ("Enterprise value por multiplos", "=Multiplos!G16", "Peso", "=Inputs!B38"),
        ("Valor patrimonial + intangiveis", "=Patrimonial!D15", "Peso", "=Inputs!B39"),
        ("Valor final estimado ponderado", "=SUMPRODUCT(B2:B4,D2:D4)/SUM(D2:D4)", "", ""),
        ("Faixa baixa (-15%)", "=B5*0.85", "", ""),
        ("Faixa alta (+15%)", "=B5*1.15", "", ""),
        ("Margem EBITDA ano 1", "=FCD!C9", "", ""),
        ("WACC", "=Inputs!B32", "", ""),
        ("Crescimento perpetuo", "=Inputs!B33", "", ""),
    ]
    for r in summary_rows:
        ws_resumo.append(list(r))
    for row in range(2, 11):
        ws_resumo.cell(row, 2).number_format = pct_fmt() if row in [9, 10, 11] else money_fmt()
        ws_resumo.cell(row, 4).number_format = pct_fmt()
    set_widths(ws_resumo, {"A": 34, "B": 20, "C": 12, "D": 12})
    style_sheet(ws_resumo)

    bar = BarChart()
    bar.title = "Valor por metodologia"
    bar.y_axis.title = "R$"
    data = Reference(ws_resumo, min_col=2, min_row=2, max_row=4)
    cats = Reference(ws_resumo, min_col=1, min_row=2, max_row=4)
    bar.add_data(data, titles_from_data=False)
    bar.set_categories(cats)
    bar.height = 7
    bar.width = 16
    ws_resumo.add_chart(bar, "F3")

    # Fontes
    ws_fontes.append(["Data de elaboracao", "16/04/2026"])
    ws_fontes.append(["Observacao", "Multiples sao inputs editaveis e devem ser atualizados com cotacoes/EV do dia antes de uma decisao vinculante."])
    ws_fontes.append(["Fonte", "Link", "Uso no modelo"])
    fonte_rows = [
        ("InfoMoney - Riachuelo 4T25", "https://www.infomoney.com.br/mercados/riachuelo-riaa3-resultados-quarto-trimestre-2025/", "Receita/EBITDA de referencia setorial"),
        ("FashionUnited - Riachuelo 2025", "https://fashionunited.com/news/business/riachuelo-parent-company-guararapes-ends-2025-with-over-100-percent-growth/2026022370741", "Resultados anuais 2025 e valor de mercado citado"),
        ("Visno Invest - Lojas Renner 2025", "https://visnoinvest.com.br/news/11946/lojas-renner-lren3-registra-lucro-de-1-5-bi-em-2025", "Receita e EBITDA anual de par"),
        ("Revista Fator Brasil - C&A 2025", "https://www.revistafatorbrasil.com.br/2026/02/26/ca-apresenta-lucro-liquido-de-r-5871-milhoes-em-2025-com-alta-de-297/", "Receita e indicadores de par"),
        ("ADVFN - Azzas 2154 4T25", "https://br.advfn.com/jornal/2026/03/azzas-2154-tem-lucro-liquido-de-r-168-milhoes-no-4t25-resultado-anual-sobe-30-5-em-2025", "EBITDA e lucro de par"),
        ("BP Money - Vulcabras 4T25", "https://bpmoney.com.br/mercado/vulcabras-vulc3-lucra-r-1588-milhoes-no-4t25-queda-de-61/", "EBITDA e receita de par"),
    ]
    for r in fonte_rows:
        ws_fontes.append(list(r))
    set_widths(ws_fontes, {"A": 34, "B": 96, "C": 42})
    style_sheet(ws_fontes)

    # Freeze panes and input highlighting
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
    input_fill = PatternFill("solid", fgColor="FFF2CC")
    for coord in ["B4:B18", "B22:B33", "B37:B39"]:
        for row in ws_inputs[coord]:
            for cell in row:
                if not str(cell.value).startswith("="):
                    cell.fill = input_fill
    for coord in ["E2:G6", "E8:G8"]:
        for row in ws_mult[coord]:
            for cell in row:
                cell.fill = input_fill
    for coord in ["B2:C9", "C14:C15"]:
        for row in ws_pat[coord]:
            for cell in row:
                if not str(cell.value).startswith("="):
                    cell.fill = input_fill

    wb.save(OUTPUT)


if __name__ == "__main__":
    main()
